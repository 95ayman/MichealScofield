import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from openpyxl import Workbook, load_workbook

# Importation de nos propres modules
import calcul_poteau
import dessin_dxf

donnees = []

def lire_excel(fichier):
    wb = load_workbook(fichier)
    ws = wb.active
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        t_sect = "Circulaire" if (row[3] is not None and float(row[3]) > 0 and (row[1] is None or float(row[1]) == 0)) else "Rectangulaire"
        data.append({
            "H": float(row[0]) if row[0] is not None else 0,
            "b": float(row[1]) if row[1] is not None else 0,
            "h": float(row[2]) if row[2] is not None else 0,
            "diam_section": float(row[3]) if row[3] is not None else 0,
            "NG": float(row[4]) if row[4] is not None else 0,
            "NQ": float(row[5]) if row[5] is not None else 0,
            "fck": float(combo_beton.get()),
            "fyk": float(combo_acier.get()),
            "type_section": t_sect
        })
    return data

def exporter_excel(resultats):
    wb = Workbook()
    ws = wb.active
    ws.append([
        "H", "b", "h", "diam_section", "NG", "NQ", "NEd", "lambda", "Ac", "fcd", "fyd", "NcRd",
        "As_calc", "As_min", "As_req", "As_fournie", "Nb_barres", "Choix_Armature", "cnom_mm", 
        "NRd", "Verification", "Type_section", "s_cl_tmax", "phi_min"
    ])
    for r in resultats:
        ws.append([
            r["H"], r["b"], r["h"], r.get("diam_section", 0), r["NG"], r["NQ"],
            r["NEd"], r["lambda"], r["Ac"], r["fcd"], r["fyd"], r["NcRd"],
            r["As_calc"], r["As_min"], r["As_req"], r["As_fournie"],
            r["nb_barres"], r["choix_armature"], r["cnom"], r["NRd"], r["verification"],
            r["type_section"], r["s_cl_tmax"], r["phi_min"]
        ])
    wb.save("resultats_poteau.xlsx")

def importer_excel():
    global donnees
    file = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
    if file:
        donnees = lire_excel(file)
        messagebox.showinfo("OK", "Excel chargé")
        calculer()

def masquer_reveler_champs(event=None):
    type_section = combo_section.get()
    if type_section == "Rectangulaire":
        lbl_b.grid(row=2, column=0, sticky="w", pady=2)
        entry_b.grid(row=2, column=1, pady=2, padx=5)
        lbl_h.grid(row=3, column=0, sticky="w", pady=2)
        entry_h.grid(row=3, column=1, pady=2, padx=5)
        lbl_diam_sect.grid_remove()
        entry_diam_section.grid_remove()
    else:
        lbl_b.grid_remove()
        entry_b.grid_remove()
        lbl_h.grid_remove()
        entry_h.grid_remove()
        lbl_diam_sect.grid(row=2, column=0, sticky="w", pady=2)
        entry_diam_section.grid(row=2, column=1, pady=2, padx=5)

def calculer():
    try:
        global donnees
        resultats = []
        type_section = combo_section.get()
        label_type.config(text=f"Type section : {type_section}")

        if donnees:
            data_source = donnees
        else:
            if type_section == "Rectangulaire":
                b_val, h_val, diam_sect_val = float(entry_b.get()), float(entry_h.get()), 0.0
            else:
                b_val, h_val, diam_sect_val = 0.0, 0.0, float(entry_diam_section.get())

            data_source = [{
                "H": float(entry_H.get()), "b": b_val, "h": h_val, "diam_section": diam_sect_val,
                "NG": float(entry_NG.get()), "NQ": float(entry_NQ.get()),
                "fck": float(combo_beton.get()), "fyk": float(combo_acier.get()),
                "type_section": type_section
            }]

        diam_etrier = float(entry_etrier.get())

        for p in data_source:
            try:
                # Appel au module de calcul
                r = calcul_poteau.calculer_ferraillage(p, diam_etrier)
                resultats.append(r)
            except ValueError as e:
                messagebox.showerror("Erreur de calcul", str(e))
                return

        exporter_excel(resultats)

        # Mise à jour de l'interface avec le dernier résultat calculé
        if resultats:
            r = resultats[-1] # On affiche les résultats du dernier poteau analysé
            label_verification_geo.config(text="Vérification géométrique : OK", fg="green")
            label_Ac.config(text=f"Ac = {r['Ac']:.2f} mm²")
            label_lambda.config(text=f"λ = {r['lambda']:.2f}")
            label_NEd.config(text=f"N_Ed = {r['NEd']:.2f} kN")
            label_fcd.config(text=f"fcd = {r['fcd']:.2f} MPa")
            label_fyd.config(text=f"fyd = {r['fyd']:.2f} MPa")
            label_NcRd.config(text=f"NcRd = {r['NcRd'] / 1000:.2f} kN")
            label_Ascalc.config(text=f"As calc = {r['As_calc']:.2f} mm²")
            label_Asmin.config(text=f"As min = {r['As_min']:.2f} mm²")
            label_Asreq.config(text=f"As req = {r['As_req']:.2f} mm²")
            label_Asfournie.config(text=f"As fournie = {r['As_fournie']:.2f} mm²")
            label_nbbarres.config(text=f"Armature choisie : {r['choix_armature']}")
            label_enrobage.config(text=f"Enrobage nominal c_nom = {r['cnom']} mm")
            label_NRd.config(text=f"NRd = {r['NRd']:.2f} kN")

            color_verif = "green" if r['verification'] == "OK" else "red"
            label_verification.config(text=f"Vérif = {r['verification']}", fg=color_verif)
            label_etrier.config(text=f"Étrier OK : Ømin = {r['phi_min']:.1f} mm")
            label_espacement.config(text=f"Espacement max étriers = {r['s_cl_tmax']:.0f} mm")

            # Déclenchement au module CAO pour le premier poteau calculé
            dessin_dxf.generer_dessin_dxf(resultats[0], "poteau_BA_rendu.dxf")

        donnees = []
        messagebox.showinfo("Succès", "Calcul terminé, export Excel et dessin DXF générés avec succès !")

    except Exception as e:
        messagebox.showerror("Erreur système", str(e))

def reset():
    for e in [entry_H, entry_b, entry_h, entry_diam_section, entry_NG, entry_NQ, entry_etrier]:
        e.delete(0, tk.END)
    combo_beton.set("25")
    combo_acier.set("500")
    combo_section.set("Rectangulaire")
    masquer_reveler_champs()

# --- INITIALISATION DE L'INTERFACE TKINTER ---
fenetre = tk.Tk()
fenetre.title("Poteau BA - Projet complet")
fenetre.geometry("850x570")

frame_actions = tk.LabelFrame(fenetre, text="Actions")
frame_actions.pack(fill="x", padx=10, pady=5)

tk.Button(frame_actions, text="Importer Excel", command=importer_excel).pack(side="left", padx=10, pady=5)
tk.Button(frame_actions, text="Calculer", command=calculer).pack(side="left", padx=10, pady=5)
tk.Button(frame_actions, text="Reset", command=reset).pack(side="left", padx=10, pady=5)

frame_principal = tk.Frame(fenetre)
frame_principal.pack(fill="both", expand=True, padx=10, pady=5)

frame_entrees = tk.LabelFrame(frame_principal, text="Entrées - Données du poteau")
frame_entrees.pack(side="left", fill="both", expand=True, padx=5, pady=5)

tk.Label(frame_entrees, text="Type de section").grid(row=0, column=0, sticky="w", pady=2)
combo_section = ttk.Combobox(frame_entrees, values=["Rectangulaire", "Circulaire"], state="readonly")
combo_section.grid(row=0, column=1, pady=2, padx=5)
combo_section.set("Rectangulaire")
combo_section.bind("<<ComboboxSelected>>", masquer_reveler_champs)

tk.Label(frame_entrees, text="H (m) - Hauteur du poteau").grid(row=1, column=0, sticky="w", pady=2)
entry_H = tk.Entry(frame_entrees)
entry_H.grid(row=1, column=1, pady=2, padx=5)

lbl_b = tk.Label(frame_entrees, text="b (cm) - Largeur de la section")
entry_b = tk.Entry(frame_entrees)
lbl_h = tk.Label(frame_entrees, text="h (cm) - Hauteur de la section")
entry_h = tk.Entry(frame_entrees)
lbl_diam_sect = tk.Label(frame_entrees, text="Diamètre section poteau (cm)")
entry_diam_section = tk.Entry(frame_entrees)

masquer_reveler_champs()

tk.Label(frame_entrees, text="N_G (kN) - Charge permanente").grid(row=4, column=0, sticky="w", pady=2)
entry_NG = tk.Entry(frame_entrees)
entry_NG.grid(row=4, column=1, pady=2, padx=5)

tk.Label(frame_entrees, text="N_Q (kN) - Charge d'exploitation").grid(row=5, column=0, sticky="w", pady=2)
entry_NQ = tk.Entry(frame_entrees)
entry_NQ.grid(row=5, column=1, pady=2, padx=5)

tk.Label(frame_entrees, text="fck (MPa) - Résistance béton").grid(row=6, column=0, sticky="w", pady=2)
combo_beton = ttk.Combobox(frame_entrees, values=[20, 25, 30, 35], state="readonly")
combo_beton.grid(row=6, column=1, pady=2, padx=5)
combo_beton.set("25")

tk.Label(frame_entrees, text="fyk (MPa) - Résistance acier").grid(row=7, column=0, sticky="w", pady=2)
combo_acier = ttk.Combobox(frame_entrees, values=[400, 500], state="readonly")
combo_acier.grid(row=7, column=1, pady=2, padx=5)
combo_acier.set("500")

tk.Label(frame_entrees, text="Ø étrier (mm)").grid(row=8, column=0, sticky="w", pady=2)
entry_etrier = tk.Entry(frame_entrees)
entry_etrier.grid(row=8, column=1, pady=2, padx=5)

res = tk.LabelFrame(frame_principal, text="Résultats")
res.pack(side="right", fill="both", expand=True, padx=5, pady=5)

font_res = ("Helvetica", 9)
font_bold = ("Helvetica", 10, "bold")

label_verification_geo = tk.Label(res, font=font_bold)
label_verification_geo.pack(anchor="w", padx=10, pady=2)
label_type = tk.Label(res, font=font_res)
label_type.pack(anchor="w", padx=10)
label_Ac = tk.Label(res, font=font_res)
label_Ac.pack(anchor="w", padx=10)
label_lambda = tk.Label(res, font=font_res)
label_lambda.pack(anchor="w", padx=10)
label_NEd = tk.Label(res, font=font_res)
label_NEd.pack(anchor="w", padx=10)
label_fcd = tk.Label(res, font=font_res)
label_fcd.pack(anchor="w", padx=10)
label_fyd = tk.Label(res, font=font_res)
label_fyd.pack(anchor="w", padx=10)
label_NcRd = tk.Label(res, font=font_res)
label_NcRd.pack(anchor="w", padx=10)
label_Ascalc = tk.Label(res, font=font_res)
label_Ascalc.pack(anchor="w", padx=10)
label_Asmin = tk.Label(res, font=font_res)
label_Asmin.pack(anchor="w", padx=10)
label_Asreq = tk.Label(res, font=font_res)
label_Asreq.pack(anchor="w", padx=10)
label_Asfournie = tk.Label(res, font=font_res)
label_Asfournie.pack(anchor="w", padx=10)

label_nbbarres = tk.Label(res, font=font_bold, fg="blue")
label_nbbarres.pack(anchor="w", padx=10, pady=2)
label_enrobage = tk.Label(res, font=font_bold, fg="darkgreen")
label_enrobage.pack(anchor="w", padx=10, pady=2)

label_NRd = tk.Label(res, font=font_res)
label_NRd.pack(anchor="w", padx=10)
label_verification = tk.Label(res, font=font_bold)
label_verification.pack(anchor="w", padx=10, pady=2)
label_etrier = tk.Label(res, font=font_res)
label_etrier.pack(anchor="w", padx=10)
label_espacement = tk.Label(res, font=font_res)
label_espacement.pack(anchor="w", padx=10)

fenetre.mainloop()